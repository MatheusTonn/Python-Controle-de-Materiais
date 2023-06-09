import tkinter as tk
from tkinter import ttk
import datetime as dt
import openpyxl as excel
import os
import tkinter.messagebox as messagebox


lista_tipos = ["Galão", "Caixa", "Saco", "Unidade"]
conjunto_codigos = set()

if os.path.exists('Controle de Materiais.xlsx'):
    #Carrega o arquivo existente
    book = excel.load_workbook('Controle de Materiais.xlsx')

    #Verifica se a planilha 'Banco de Dados' já existe no arquivo
    if 'Banco de Dados' in book.sheetnames:
        #Acessa a planilha existente
        banco_de_dados = book['Banco de Dados'] 
    else:
        #Cria uma nova planilha 'Banco de Dados'
        banco_de_dados = book.create_sheet('Banco de Dados')
else:
    #Cria um novo arquivo e uma nova planilha 'Banco de Dados'
    book = excel.Workbook()
    banco_de_dados = book.create_sheet('Banco de Dados')

janela = tk.Tk()
#Criação da Função
def inserir_codigo():

    if entry_descricao.get() == "" or entry_quant.get() == "" or combobox_selecionar_tipo.get() == "":
        messagebox.showerror("Erro", "Os campos não foram preenchidos corretamente!")
    else: 
        #Pega as informações dos campos de inserção
        descricao = entry_descricao.get()
        tipo = combobox_selecionar_tipo.get()
        quantidade = int(entry_quant.get())
        #Pega a data e hora de inserção do material
        data_criacao = dt.datetime.now()
        data_criacao = data_criacao.strftime("%d/%m/%Y %H:%M")
        #Insere o material da Lista de códigos
        codigo = len(conjunto_codigos)+1
        conjunto_codigos.add((descricao, tipo, quantidade, data_criacao))    
        banco_de_dados.append((descricao, tipo, quantidade, data_criacao))
        #Campo Label "Mensagem de Sucesso"
        label_mensagem_sucesso.config(text="CÓDIGO INSERIDO")
        label_mensagem_sucesso.grid(row=6, column=0, sticky='nswe', columnspan=4)
        #Agendar a remoção dos campos
        janela.after(1000, remover_campos)

        entry_descricao.focus_set()

def remover_campos():
    label_mensagem_sucesso.grid_remove()
    entry_descricao.delete(0, tk.END)
    combobox_selecionar_tipo.set('')
    entry_quant.delete(0, tk.END)

def fechar_janela():
    if entry_descricao.get() or combobox_selecionar_tipo.get() or entry_quant.get():
        resposta = messagebox.askyesno("Confirmação", "Deseja cancelar a operação e fechar a janela?")
        if resposta:
            janela.destroy()
            #Salva planilha
            book.save('Controle de Materiais.xlsx')
    else:
        janela.destroy()
        #Salva planilha
        book.save('Controle de Materiais.xlsx')

#Titulo da Janela
janela.title('Ferramenta de Cadastro de Materiais')

#Campo de Label "Descrição do Material"
label_descricao = tk.Label(text="Descrição do Material")
label_descricao.grid(row=1, column=0, padx=10, pady=10, sticky='nswe', columnspan=4)

#Campo de Entry "Descrição do Material"
entry_descricao = tk.Entry() 
entry_descricao.grid(row=2, column=0, padx=10, pady=10, sticky='nswe', columnspan=4)

#Campo de Entry "Tipo de Unidade do Material"
label_tipo_unidade = tk.Label(text="Tipo da Unidade do Material")
label_tipo_unidade.grid(row=3, column=0, padx=10, pady=10, sticky='nswe', columnspan=2)

#Input Box de seleção de dados com base na Lista: lista_tipos
combobox_selecionar_tipo = ttk.Combobox(values=lista_tipos)
combobox_selecionar_tipo.grid(row=3, column=2, padx=10, pady=10, sticky='nswe', columnspan=2)

#Campo Label "Mensagem de sucesso"
label_mensagem_sucesso = tk.Label(text="")

#Quantidade do Tipo
label_quant = tk.Label(text="Quantidade da Unidade")
label_quant.grid(row=4, column=0, padx=10, pady=10, sticky='nswe', columnspan=2)

entry_quant = tk.Entry()
entry_quant.grid(row=4, column=2, padx=10, pady=10, sticky='nswe', columnspan=2)

botao_criar_codigo = tk.Button(text="Criar código", command=inserir_codigo)
botao_criar_codigo.grid(row=5, column=0, padx=10, pady=10, sticky='nswe', columnspan=4)
#Definir ação de fechar janela
janela.protocol("WM_DELETE_WINDOW", fechar_janela)


janela.mainloop()

